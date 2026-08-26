"""Root pytest configuration for QA-Automation."""

import os
import subprocess
from datetime import datetime
from pathlib import Path

import pytest
from dotenv import load_dotenv

OUTCOMES_DIR = Path(os.environ.get("QA_OUTCOMES_DIR", "reports/outcomes"))
ALLURE_RESULTS = Path("allure-results")

_run_timestamp: str = ""
_test_results: list[dict] = []


def _fetch_deployed_branch() -> str:
    return os.environ.get("DEPLOYED_BRANCH", "unknown")


def _write_allure_environment():
    branch   = _fetch_deployed_branch()
    env      = os.environ.get("ENV", "staging")
    base_url = os.environ.get("BASE_URL", "https://staging-mercato.skoopin.net")
    ALLURE_RESULTS.mkdir(parents=True, exist_ok=True)
    (ALLURE_RESULTS / "environment.properties").write_text(
        f"Deployed.Branch={branch}\n"
        f"Environment={env}\n"
        f"Base.URL={base_url}\n"
    )


def pytest_configure(config):  # noqa: ARG001
    global _run_timestamp
    _run_timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")


def pytest_sessionstart(session):  # noqa: ARG001
    _write_allure_environment()


@pytest.hookimpl(hookwrapper=True)
def pytest_runtest_makereport(item, call):
    outcome = yield
    report = outcome.get_result()
    if report.when == "call":
        marker = item.get_closest_marker("testcase")
        if marker:
            _test_results.append({
                "component":    marker.kwargs.get("component", ""),
                "type":         marker.kwargs.get("type", ""),
                "description":  marker.kwargs.get("description", ""),
                "steps":        marker.kwargs.get("steps", ""),
                "status":       report.outcome.upper(),
                "error_detail": report.longreprtext if report.failed else "",
            })


@pytest.hookimpl(trylast=True)  # run after junitxml writes reports/junit.xml, which Xray reporting reads
def pytest_sessionfinish(session, exitstatus):  # noqa: ARG001
    if not _run_timestamp:
        return

    run_dir = OUTCOMES_DIR / f"run_{_run_timestamp}"
    run_dir.mkdir(parents=True, exist_ok=True)

    # Allure report
    subprocess.run(
        ["allure", "generate", str(ALLURE_RESULTS), "-o", str(run_dir), "--clean", "--single-file"],
        check=False,
    )

    # AI triage for failures — reads context/tickets.json built ahead of time by
    # scripts/build_context.py. Skipped (not guessed) if that context doesn't exist.
    failed_results = [r for r in _test_results if r["status"] == "FAILED"]
    if failed_results:
        import sys
        sys.path.insert(0, str(Path(__file__).parent / "scripts"))
        from failure_triage import analyze_failure, load_release_context

        release_context = load_release_context()
        if release_context is None:
            print("AI triage: context/tickets.json not found — run scripts/build_context.py "
                  "with this release's ticket keys first. Skipping AI analysis.")
        else:
            print(f"AI triage: classifying {len(failed_results)} failure(s) ...")
            for result in failed_results:
                result["ai"] = analyze_failure(result, release_context)
                print(f"AI triage: {result['description'][:60]!r} -> "
                      f"{result['ai']['verdict']} (confidence {result['ai']['confidence']:.2f})")

    # Excel report
    if _test_results:
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Test Results"

            ws.append(["Component", "Type", "Description", "Steps", "Status",
                       "AI Verdict", "Confidence", "Matched Ticket", "AI Reasoning"])
            for cell in ws[1]:
                cell.font = Font(bold=True)

            status_colors = {"PASSED": "92D050", "FAILED": "FF4C4C", "ERROR": "FFA500"}
            verdict_colors = {"bug": "FF4C4C", "suite_improvement": "FFD966", "needs_review": "D9D9D9"}

            for result in _test_results:
                ai = result.get("ai")
                ws.append([
                    result["component"],
                    result["type"],
                    result["description"],
                    result["steps"],
                    result["status"],
                    ai["verdict"] if ai else "",
                    round(ai["confidence"], 2) if ai else "",
                    ai["matched_ticket"] if ai else "",
                    ai["reasoning"] if ai else "",
                ])
                last_row = ws.max_row
                color = status_colors.get(result["status"], "FFFFFF")
                ws.cell(last_row, 5).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                ws.cell(last_row, 4).alignment = Alignment(wrap_text=True)
                if ai:
                    v_color = verdict_colors.get(ai["verdict"], "FFFFFF")
                    ws.cell(last_row, 6).fill = PatternFill(start_color=v_color, end_color=v_color, fill_type="solid")
                    ws.cell(last_row, 9).alignment = Alignment(wrap_text=True)

            ws.column_dimensions["A"].width = 15
            ws.column_dimensions["B"].width = 20
            ws.column_dimensions["C"].width = 55
            ws.column_dimensions["D"].width = 60
            ws.column_dimensions["E"].width = 12
            ws.column_dimensions["F"].width = 18
            ws.column_dimensions["G"].width = 12
            ws.column_dimensions["H"].width = 16
            ws.column_dimensions["I"].width = 60

            wb.save(str(run_dir / "test_results.xlsx"))
        except ImportError:
            pass

    # Xray Cloud — report results for tests already onboarded (never creates a Test).
    # Skipped entirely if credentials aren't set, so local dev without them is unaffected.
    if os.environ.get("XRAY_CLIENT_ID") and os.environ.get("XRAY_CLIENT_SECRET"):
        try:
            import sys
            sys.path.insert(0, str(Path(__file__).parent / "scripts"))
            from xray_common import push_execution_results
            from xray_report import build_entries

            print("Xray: matching test results against onboarded keys ...")
            keyed, skipped = build_entries()
            print(f"Xray: {len(keyed)} keyed result(s) matched, {len(skipped)} unkeyed result(s) skipped.")
            if keyed:
                build = os.environ.get("BUILD_NUMBER", "Local Execution")
                execution = push_execution_results(keyed, build)
                print(f"Xray: reported {len(keyed)} result(s) -> {execution['key']}")
            if skipped:
                print(f"Xray: {len(skipped)} test(s) ran without a key, skipped reporting — "
                      f"run scripts/xray_onboard.py to onboard them: {', '.join(skipped)}")
        except Exception as exc:
            import traceback
            print(f"Xray: reporting failed, non-fatal ({exc})")
            traceback.print_exc()
    else:
        print("Xray: XRAY_CLIENT_ID/XRAY_CLIENT_SECRET not set — skipping Xray reporting.")


load_dotenv(Path(__file__).parent / ".env")

from dashboard.fixtures.browser_fixtures import *  # noqa: F401, F403, E402
from dashboard.fixtures.auth_fixtures import *     # noqa: F401, F403, E402
from dashboard.fixtures.api_fixtures import *      # noqa: F401, F403, E402
